import time
import os
import re
import openpyxl
import pandas as pd
import numpy  as np
import seaborn as sns
import matplotlib.pyplot as plt
excelPath  = 'E:/student_data/secret/选课统计信息（热力图）/all_file'# 根据文件夹位置更改
toPath='E:/student_data/secret/选课统计信息（热力图）/to_file'
schemePlanNumberToStudentNumberDict = {}                                     # 方案计划号作为键，学号作为值
termToNum  = {}                                                              # 学期课表对应筛选条件的字典
studentWb  = {}                                                              # 学生选课表对应其年份的字典
ChiDays    = ['星期一','星期二','星期三','星期四','星期五','星期六','星期日']# 中文一周七天列表

filenames  = os.listdir(excelPath)                                                  # 获取所有的excel
filenames2 = []


for i in range(len(filenames)):
    filenames2.append(filenames[i].split('.')[0])

filenames1To2 = dict(zip(filenames,filenames2))

    
for i in filenames:
    if '春' in i:                                                                  # 获取春季学期课表
        startAndEndTime = re.findall(r'\d+', i)
        staTi = startAndEndTime[0]
        endTi = startAndEndTime[1]
        num   = '20' + str(staTi) + '-' + '20' + str(endTi) + '-' + '2' + '-' + '1'# 获取相应筛选条件
        termToNum.setdefault(i,num)  
        os.mkdir(excelPath + '/' + filenames1To2[i])                                       # 对每个春季学期创建一个文件夹
    elif '秋' in i:                                                                # 获取秋季学期课表
        startAndEndTime = re.findall(r'\d+', i)
        staTi = startAndEndTime[0]
        endTi = startAndEndTime[1]
        num   = '20' + str(staTi) + '-' + '20' + str(endTi) + '-' + '1' + '-' + '1'# 获取相应筛选条件
        termToNum.setdefault(i,num)
        os.mkdir(excelPath + '/' + filenames1To2[i])                                       # 对每个秋季学期创建一个文件夹
    else:                                                                          # 获取学生选课表
        startAndEndTime = re.findall(r'\d+', i)
        staTi = startAndEndTime[0]
        endTi = startAndEndTime[1]
        allTi = []
        for j in range(int(staTi),int(endTi) + 1):                                 # 获取对应年份列表
            allTi.append(str(j))
            studentWb.setdefault(i,allTi)

final_df=pd.DataFrame()

for i in termToNum.keys():
    print(str(i)+"-------------------")
    df1 = pd.DataFrame(pd.read_excel(excelPath + '/' + i,sheet_name = 'Sheet1'))                                               # 打开学期选课表
    num = re.findall(r'\d+', i)[0]
    for j in studentWb.values():
        if num in j:
            name = list(studentWb.keys())[list(studentWb.values()).index(j)]
            df2 = pd.DataFrame(pd.read_excel(excelPath + '/' + name,sheet_name = 'Sheet1'))                                    # 打开相应年份的学生选课表
            df2 = df2.loc[df2['开课学期'] == termToNum[i]]                                                                     # 对学生选课表进行年份的条件筛选
            vlookupWb = pd.merge(df1,df2.loc[ : , ['课程号','课序号','年级','方案计划号','学号']],how='left',on = ['课程号','课序号'])# 进行两表合并
            vlookupWb.dropna(subset = ['年级'],inplace = True)
            vlookupWb.dropna(subset = ['方案计划号'],inplace = True)
            vlookupWb.dropna(subset = ['上课周次'],inplace = True)                                                             # 年级、方案计划号和上课周次删除空值所在行
            vlookupWb.dropna(subset = ['学号'],inplace = True)
            for k in ChiDays:
                vlookupWb[k].fillna(0,inplace = True)                                                                          # 每天上课时间空值替换成0
            writer = pd.ExcelWriter(excelPath + '/' + filenames1To2[i] + '/' + filenames1To2[i] + '.xlsx')
            vlookupWb.to_excel(writer,index=False)
            writer.save()                                                                                                      # 保存新表



            # 第二部分：生成热力图
                                    
                       
            data    = {}                                                             # 总数据的字典
            OrderOfRetrieval  = ['年级',      
                                '方案计划号',
                                '上课周次' ]                                         # 中文检索的列表
            engNameOfRet      = ['grade',
                                'SchemePlanN',
                                'numOfWeeks']                                        # 英文检索的列表
            chiToEngRet       = dict(zip(OrderOfRetrieval,engNameOfRet))             # 中英文检索对应字典
            retNo             = []    
            realCol           = []    
            days    = ['mon','tue','wed','thu','fri','sat','sun']                    # 英文一周7天列表
            classes = list(range(1,14))                                              # 一天13小节
            weeks   = list(range(1,23))                                              # 一个学期22周(最大周数，可能不足)
            
            
            thermodynamicPath = excelPath + '/' + filenames1To2[i] + '/热力图'       # 热力图保存路径
            termname=filenames1To2[i]
            os.mkdir(thermodynamicPath)                                              # 创建热力图保存路径
            excelPath2        = excelPath + '/' + filenames1To2[i] + '/' + filenames1To2[i] + '.xlsx'# 生成的excel工作簿路径
            
            
            wb    = openpyxl.load_workbook(excelPath2)# 打开生成的excel工作簿
            sheet = wb['Sheet1']                      # 打开工作表


            def CourseStatistics():                                                                                                         # 统计课程的函数(可供以后修改)
                for everyday in range(len(days)):
                    classNo = sheet.cell(row    = rowN,                                                                                     
                                         column = mondayCol+everyday).value                                                                 # 遍历一周课程情况                                                 
                    firAndLasC = re.findall(r'\d+', str(classNo))
                    if   len(firAndLasC) == 1:                                                                                              # 一周一次课 每次一个小节课程
                        if int(classNo)       != 0:
                            data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(k)][days[everyday]].setdefault(int(classNo),0)
                            data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(k)][days[everyday]][int(classNo)] += 1# 获取上课次数
                    elif len(firAndLasC) == 2:                                                                                              # 一周一次课 每次多个小节课程
                        for l in range(int(firAndLasC[0]),int(firAndLasC[1]) + 1 ):
                            data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(k)][days[everyday]].setdefault(l,0)
                            data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(k)][days[everyday]][l] += 1
                    elif len(firAndLasC) == 4:                                                                                              # 一周两次课 每次多个小节课程
                        for m in range(int(firAndLasC[0]),int(firAndLasC[1]) + 1 ):
                            data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(k)][days[everyday]].setdefault(m,0)
                            data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(k)][days[everyday]][m] += 1
                        for n in range(int(firAndLasC[2]),int(firAndLasC[3]) + 1 ):
                            data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(k)][days[everyday]].setdefault(n,0)
                            data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(k)][days[everyday]][n] += 1
            
            
            def modifyData():                                                                       # 修改data
                global data
                data                  = dict(sorted(data.items(), key=lambda x:x, reverse=True))    # 年份从大到小
                for i in list(data.keys()):
                    data[i]           = dict(sorted(data[i].items(), key=lambda x:x, reverse=False))# 方案计划号从小到大
                firstKetList          = list(data.keys())
                for i in firstKetList:
                    secondKeyList     = list(data[i].keys())
                    for j in secondKeyList:
                        for k in weeks:
                            myData    = data[i][j]['week'+str(k)]
                            dataframe = pd.DataFrame(myData,columns=days,index=classes,dtype = np.int8)
                            if np.all(dataframe == 0) == True:                                      # 检查是否一周均无课
                                data[i][j].pop('week'+str(k))                                       # 删除无课一周的数据
            
            
            for colN in range(1,sheet.max_column+1):
                cellOfFirRow     = sheet.cell(row = 1,column = colN).value                                                       # 检查工作表第一行
                eleOfRealCol     = colN
                if cellOfFirRow == '星期一':
                    mondayCol    = colN                                                                                          # 获取 列'星期一' 对应列数                
                if cellOfFirRow == '学号':
                    studentNumberCol = colN
                if cellOfFirRow == '方案计划号':
                    schemePlanNumberCol = colN
                    
                if cellOfFirRow in OrderOfRetrieval:                
                    retNo.append(chiToEngRet[cellOfFirRow])                                                                      # 获取检索，放到retNo列表里
                    realCol.append(eleOfRealCol)                                                                                 # 获取检索的对应列数，放到realCol列表里
            retNoAndRealColDict  = dict(zip(retNo,realCol))                                                                      # 检索与列数对应的字典


            for i in range(2,sheet.max_row + 1):
                studentNumber = sheet.cell(row = i, column = studentNumberCol).value
                schemePlanNumber = sheet.cell(row = i, column = schemePlanNumberCol).value
                schemePlanNumberToStudentNumberDict.setdefault(studentNumber,schemePlanNumber)
            dict2 = {}
            for j in list(schemePlanNumberToStudentNumberDict.values()):
                dict2.setdefault(j,0)
                dict2[j] += 1
            schemePlanNumberToStudentNumberDict = dict2                
            
            
            for rowN in range(2,sheet.max_row + 1):
                for ret in retNo:
                    globals()[ret] = sheet.cell(row    = rowN,column = retNoAndRealColDict[ret]).value                           # 遍历检索对应列的值                                                                                                                                                                                                            
                data.setdefault(globals()[engNameOfRet[1 -1]],{})                                                                # 获取字典第一个键，‘年级’
                data[globals()[engNameOfRet[1 -1]]].setdefault(globals()[engNameOfRet[2 -1]],{})                                 # 获取字典第二个键，‘方案计划号’
                for i in weeks:
                    data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]].setdefault('week'+ str(i),{})             # 设置字典第三个键，周次
                    for j in days:
                        data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(i)].setdefault(j,{})      # 设置字典第四个键，天次
                        for k in classes:
                            data[globals()[engNameOfRet[1 -1]]][globals()[engNameOfRet[2 -1]]]['week'+ str(i)][j].setdefault(k,0)# 设置字典第五个键，课次
            
                
            
                # '/'表示有两种上课周次，这里两种周次均一次算入；'-'表示周次是否连续，'单' '双'分别表示单周、双周                

                if ('单' in numOfWeeks) and ('/' not in numOfWeeks):                                # 一种周次 单周上课
                    firAndLasW = re.findall(r'\d+', str(numOfWeeks))                             
                    firW       = firAndLasW[0]
                    lasW       = firAndLasW[1]                                                      # 获取起始和结束周次
                    for k in range(int(firW),int(lasW) + 1,2):                                      # 隔周计算
                        CourseStatistics()
            
                if ('双' in numOfWeeks) and ('/' not in numOfWeeks):                                # 一种周次 双周上课
                    firAndLasW = re.findall(r'\d+', str(numOfWeeks))
                    firW       = firAndLasW[0]
                    lasW       = firAndLasW[1]
                    for k in range(int(firW),int(lasW) + 1,2):
                        CourseStatistics()
            
                if ('单' not in numOfWeeks) and ('-' in numOfWeeks) and ('/' not in numOfWeeks):    # 一种周次 每周上课                
                    firAndLasW = re.findall(r'\d+', str(numOfWeeks))
                    firW       = firAndLasW[0]
                    lasW       = firAndLasW[1]
                    for k in range(int(firW),int(lasW) + 1,1):                                      # 每周计算
                        CourseStatistics()
                                            
                if ('单' not in numOfWeeks) and ('-' not in numOfWeeks) and ('/' not in numOfWeeks):# 一种周次 间断周次               
                    firAndLasW = re.findall(r'\d+', str(numOfWeeks))                                # 获取所有间断周次
                    for j in range(len(firAndLasW)):
                        k = firAndLasW[j]                                                           # 每周计算
                        CourseStatistics()
                    
                if ('单' in numOfWeeks) and ('/' in numOfWeeks):                                    # 两种周次 前一种每周 后一种单周                                                               
                    firAndLasW = re.findall(r'\d+', str(numOfWeeks))
                    firW       = firAndLasW[0]
                    lasW       = firAndLasW[1]
                    for k in range(int(firW),int(lasW) + 1,1):
                        CourseStatistics()
                    firW       = firAndLasW[2]
                    lasW       = firAndLasW[3]
                    for k in range(int(firW),int(lasW) + 1,2):
                        CourseStatistics()
                            
                if ('双' in numOfWeeks) and ('/' in numOfWeeks):                                    # 两种周次 前一种每周 后一种双周                                                               
                    firAndLasW = re.findall(r'\d+', str(numOfWeeks))
                    firW       = firAndLasW[0]
                    lasW       = firAndLasW[1]
                    for k in range(int(firW),int(lasW) + 1,1):
                        CourseStatistics()
                    firW       = firAndLasW[2]
                    lasW       = firAndLasW[3]
                    for k in range(int(firW),int(lasW) + 1,2):
                        CourseStatistics()
                    
                if ('单' not in numOfWeeks) and ('-' in numOfWeeks) and ('/' in numOfWeeks):        # 两种周次 均为每周             
                    firAndLasW = re.findall(r'\d+', str(numOfWeeks))
                    firW       = firAndLasW[0]
                    lasW       = firAndLasW[1]
                    for k in range(int(firW),int(lasW) + 1,1):
                        CourseStatistics()
                    firW       = firAndLasW[2]
                    lasW       = firAndLasW[3]
                    for k in range(int(firW),int(lasW) + 1,1):
                        CourseStatistics()
            
            
            modifyData()           
            firstKetList          = list(data.keys())
            for i in firstKetList:                                                                            
                secondKeyList     = list(data[i].keys())
                for j in secondKeyList:                                                                       
                    data[i][j].setdefault('all',{})
                    for k in days:
                        data[i][j]['all'].setdefault(k,{})
                        for l in classes:
                            data[i][j]['all'][k].setdefault(l,0)# 周次里加上一个总周次


            firstKetList          = list(data.keys())
            for i in firstKetList:
                secondKeyList     = list(data[i].keys())
                for j in secondKeyList:
                    myList        = list(data[i][j].keys())
                    myList.remove('all')
                    thirdKeyList  = myList
                    for k in thirdKeyList:
                        fourthKeyList = list(data[i][j][k].keys())
                        for l in fourthKeyList:
                            fifthKeyList = list(data[i][j][k][l].keys())
                            for m in fifthKeyList:
                                data[i][j]['all'][l][m] += int(data[i][j][k][l][m])# 统计总周次的课程次数

            firstKetList          = list(data.keys())
            for i in firstKetList:
                secondKeyList     = list(data[i].keys())
                for j in secondKeyList:
                    myData        = data[i][j]['all']
                    dataframe = pd.DataFrame(myData,columns=days,index=classes)
                    #f = lambda num : round(num/schemePlanNumberToStudentNumberDict[l])
                    #dataframe = dataframe.applymap(f)
                    filename=str(termname)+" "+str(i)+" "+str(j)+".csv"
                    print(filename)
                    #我设置了一个toPath
                    dataframe.to_csv(toPath + '/'+filename)
                    #sns.heatmap(dataframe,cmap="RdPu", annot=True, cbar=True, fmt='.20g')# 生成热力图
                    #plt.savefig(thermodynamicPath + '/' + str(i)+ str(j))
                    #plt.clf()

                                                                                  
print('\n已完成')
            
            
    


